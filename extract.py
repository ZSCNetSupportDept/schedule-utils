#!/usr/bin/env python3

import argparse
from collections import namedtuple
from pprint import pprint

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet import Worksheet

from blocks import Block
from formatter_sql import format_sql

LEADER_ROW = 32
BLOCK_ROW_MAP = {  # block -> row
    Block.FX: 2,
    Block.BM: 7,
    Block.DM: 12,
    Block.QT: 17,
    Block.XH: 22,
    Block.ZH: 27
}


def extract_staff_name(name: str) -> str:
    return name.split('-')[0].strip()


ScheduleOfWeek = namedtuple('ScheduleOfWeek', ['leader', 'staffs'])


def extract(filename):
    wb: Workbook = load_workbook(filename=filename)
    ws: Worksheet = wb.active
    schedule = {}
    for week in range(1, 7):
        column = chr(ord('A') + week)
        leader = extract_staff_name(ws[f"{column}{LEADER_ROW}"].value)
        staffs = {}
        for block, rowIndex in BLOCK_ROW_MAP.items():
            staffs[block] = []
            for i in range(5):
                cell = ws[f"{column}{rowIndex + i}"]
                if cell.value:
                    staffs[block].append(extract_staff_name(cell.value))
        schedule[week] = ScheduleOfWeek(leader, staffs)
    return schedule


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Parse schedule excel into something useful.')
    parser.add_argument('-f', '--format', type=str, default='print',
                        help='the output format')
    parser.add_argument('-t', '--table', type=str, default='operators',
                        help='the table of sql_update output')
    parser.add_argument('input', metavar='INPUT',
                        help='the input file')
    args = parser.parse_args()
    schedule = extract(args.input)
    {
        'print': pprint,
        'sql_update': lambda schedule: format_sql(schedule, table=args.table),
    }[args.format](schedule)
